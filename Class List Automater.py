import pandas as pd
import numpy as np
import csv
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, Font, Border, Side, PatternFill, Alignment


def get_names(question_df):
    '''
    Get Names of People in PGN 

    @parameter question_df: The Dataframe gathered from the Questionnaire

    @return name_list: The List of the names Acquired from the Dataframe
    '''

    name_list = None

    for col in question_df.columns:
        
        if "name" in col.lower():

            name_list = question_df[col]
            break

    if len(name_list) == 0:
        print("Error: Ensure that Name Column in Questionnaire.")
    
    return name_list



def get_classes(name_list, question_df):
    '''
    Gets the Classes of our PGN Members

    @parameter name_list: The list of names for PGN members
    @parameter question_df: The Dataframe gathered from the Questionnaire

    @return class_dict: The classes mapped to the respective PGN members that are taking the class
    '''

    #Create a DataFrame of Just the Class List
    class_df = pd.DataFrame()

    for col in question_df.columns:

        if "class" in col.lower():
            
            class_df[col] = question_df[col]


    #Create a dictionary that maps our members to their respective classes
    name_dict = {}

    for i, name in enumerate(name_list):
        name_dict[name] = [class_df.loc[i,"Class #1:"], class_df.loc[i,"Class #2:"], class_df.loc[i,"Class #3:"],\
        class_df.loc[i,"Class #4:"], class_df.loc[i,"Class #5:"], class_df.loc[i,"Class #6:"]]
    
    #print(name_dict)


    #Validate the Classes - Invalid Classes will be Ignored
    classes_dict = dict()

    #Iterate through the classes (rows) in our class dataframe
    for i in range(len(class_df)):
        
        #Iterate through each class (cell) of the classes (row)
        for val in class_df.loc[i]:

            #Validate to see if our classes are alright to read in
            ok_class = class_validator(val)

            if ok_class == True:

                val = val.strip()

                #Some classes won't have a space b/w the letters and numbers of the class code. We will 
                #submit an extra space if one doesn't exist
                if " " not in val:
                    char = 0
                    while val[char].isalpha():

                        char += 1


                    val = val[0:char] + " " + val[char:]

                #Now we will uppercase every character in the string so that all course codes will be the same
                #despite case differences
                val = val.upper()

                #Now we can add the filtered course code to the set
                classes_dict[val] = []

    for name, classes in name_dict.items():
        
        #Iterate through the classes for each person
        for val in classes:

            #Validate to see if our classes are alright to read in
            ok_class = class_validator(val)

            if ok_class == True:

                val = val.strip()

                #Some classes won't have a space b/w the letters and numbers of the class code. We will 
                #submit an extra space if one doesn't exist
                if " " not in val:
                    char = 0
                    while val[char].isalpha():

                        char += 1


                    val = val[0:char] + " " + val[char:]

                #Now we will uppercase every character in the string so that all course codes will be the same
                #despite case differences
                val = val.upper()

                #Now we can add the filtered course code to the set
                
                classes_dict[val].append(name)

    return classes_dict


def class_validator(class_check):
    '''
    Checks to see if our person entered in a valid class into the system

    @param class_check: the class in question
    @return Boolean: the truth value of the class validation
    '''


    #Check to see if dataframe value is NaN (Not a Number)
    try:
        float_check = float(class_check)

    except:
        float_check = 1000
    

    if np.isnan(float_check):
        return False

    #Strip any extra accidental whitespace
    class_check = class_check.strip()

    #This filters out any class codes that are not submitted with a starting letter and does not have a numeric ending
    #or some unique ending numeric class codes
    if not(class_check[0].isalpha()) or (not(class_check[-1].isnumeric()) and not(class_check[-1].upper() == "L") \
        and not(class_check[-1].upper() == "H") and not(class_check[-1].upper() == "B") \
            and not(class_check[-1].upper() == "A")):
        return False

    else:
        return True


def style_spreadsheet():

    #Open File for formatting
    wb = load_workbook(filename = 'output.xlsx')

    #Open File for reading information
    scholarships = pd.read_excel("output.xlsx")

    #Determine number of rows that we must format
    num_scholarships = len(scholarships)

    #Give a desintation to where the styled notebook will go
    dest_filename = 'Class List.xlsx'

    #Start formatting
    ws1 = wb.active
    ws1.delete_rows(idx=1) #Delete id row

    #Add a style to the column labels
    col_label_style = NamedStyle(name="column heading")
    col_label_style.font = Font(bold=True, size=12)
    col_label_style.fill = PatternFill("solid", fgColor="FFD700")
    col_label_style.alignment = Alignment(horizontal="center", vertical="center")
    bd = Side(style='thick', color="000000") #black
    col_label_style.border = Border(left=bd, top=bd, right=bd, bottom=bd)

    wb.add_named_style(col_label_style)
 
    for cell in ws1["A"]:
        cell.style = col_label_style

    wb.save(filename = dest_filename)


if __name__ == "__main__":
        
    input_file = input("Enter the Questionnaire File Name: ")

    question_df = pd.read_csv(input_file)

    #Get names from our questionnaire
    name_list = get_names(question_df)

    class_dict = get_classes(name_list, question_df)

    final_class_df = pd.DataFrame({ key:pd.Series(value) for key, value in class_dict.items() })

    final_class_df = final_class_df.T

    final_class_df.to_excel("output.xlsx")

    style_spreadsheet()

    

    
